"""
Pasa a la base la planilla de vencimientos de licencias.

La planilla (VENCIMIENTO DE LICENCIAS, en Drive) tiene una hoja por
sucursal, y en cada una una fila por chofer con el dominio que maneja al
lado. Las columnas de días que faltan son fórmulas: no se importan, las
calcula la base sola contra la fecha de hoy.

    A CHOFER   B DOMINIO
    C V.T.V. alta                 D V.T.V. vencimiento
    E Licencia municipal alta     F ... vencimiento
    G Licencia profesional alta   H ... vencimiento
    I Matafuegos alta             J ... vencimiento
    O Observaciones

Uso:
    # bajar la planilla como .xlsx desde Drive (Archivo → Descargar)
    export SUPABASE_DB_URL='...'
    python importar_vencimientos.py "VENCIMIENTO DE LICENCIAS.xlsx"

    # para mirar qué haría sin escribir nada:
    python importar_vencimientos.py archivo.xlsx --probar

Se puede correr las veces que haga falta: una fecha ya cargada no se
duplica, y una fecha nueva entra como renovación arriba de la anterior.
"""
import os
import sys
import datetime

import openpyxl
import psycopg
from psycopg.rows import dict_row

# La planilla nombra las sucursales enteras y el maestro de unidades las
# abrevia. Sin traducirlas, el panel muestra "LARGA DISTANCIA" y "LAD" como
# si fueran dos lugares distintos, y cada uno con la mitad de los datos.
SUCURSALES = {
    "LARGA DISTANCIA": "LAD",
    "TUCUMAN": "TUC",
    "BUENOS AIRES": "BUE",
    "LA RIOJA": "LRJ",
    "CATAMARCA": "CAT",
    "BELEN": "BEL",
    "CORDOBA": "COR",
    "SALTA": "SAL",
    "TALLER": "TAL",
}


def sucursal_de(hoja):
    """El código con el que la sucursal se llama en el resto del sistema."""
    nombre = " ".join(str(hoja or "").upper().split())
    return SUCURSALES.get(nombre, nombre[:20])


# Las hojas que no son sucursales.
NO_SON_SUCURSALES = {"respuestas de formulario 1", "tablas_juntas", "tablas juntas"}

# Columna de alta, columna de vencimiento y tipo en la base (base 0).
COLUMNAS = [
    (2, 3, "VTV"),
    (4, 5, "Licencia municipal"),
    (6, 7, "Licencia profesional"),
    (8, 9, "Matafuegos"),
]
COL_CHOFER, COL_DOMINIO, COL_OBS = 0, 1, 14


def patente(texto):
    """AF 470 UT, af796ix y AF796IX son la misma. FLG 593 también vale."""
    return "".join(ch for ch in str(texto or "").upper() if ch.isalnum())


def nombre(texto):
    return " ".join(str(texto or "").split()).title()


def fecha(valor):
    if isinstance(valor, datetime.datetime):
        return valor.date()
    if isinstance(valor, datetime.date):
        return valor
    texto = str(valor or "").strip()
    if not texto:
        return None
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def leer(archivo):
    """Devuelve las filas de todas las hojas de sucursal."""
    wb = openpyxl.load_workbook(archivo, data_only=True)
    filas = []
    for hoja in wb.sheetnames:
        if hoja.strip().lower() in NO_SON_SUCURSALES:
            continue
        ws = wb[hoja]
        # Las dos primeras filas son el encabezado en dos pisos.
        for celdas in ws.iter_rows(min_row=3, values_only=True):
            chofer = nombre(celdas[COL_CHOFER] if len(celdas) > COL_CHOFER else None)
            pat = patente(celdas[COL_DOMINIO] if len(celdas) > COL_DOMINIO else None)
            if not chofer and not pat:
                continue
            obs = celdas[COL_OBS] if len(celdas) > COL_OBS else None
            fechas = {}
            for col_alta, col_vence, tipo in COLUMNAS:
                alta = fecha(celdas[col_alta]) if len(celdas) > col_alta else None
                vence = fecha(celdas[col_vence]) if len(celdas) > col_vence else None
                if vence:
                    fechas[tipo] = (alta, vence)
            filas.append({"sucursal": sucursal_de(hoja), "chofer": chofer,
                          "patente": pat, "observaciones": str(obs).strip() if obs else None,
                          "fechas": fechas})
    return filas


def importar(cx, filas, probar=False):
    tipos = {t["nombre"]: t for t in
             cx.execute("select * from tipos_vencimiento").fetchall()}
    unidades = {u["patente"]: u["id"] for u in
                cx.execute("select id, patente from unidades").fetchall()}

    resumen = {"unidades_nuevas": [], "personas_nuevas": 0, "cargados": 0,
               "ya_estaban": 0, "sin_fecha": 0, "sin_patente": []}

    for f in filas:
        # --- la unidad
        unidad_id = None
        if f["patente"]:
            unidad_id = unidades.get(f["patente"])
            if unidad_id and not probar:
                # La planilla sabe en qué sucursal está cada unidad; la de
                # gomería no siempre. Se completa, nunca se pisa.
                cx.execute("""update unidades set sucursal = %s
                              where id = %s and sucursal is null""",
                           (f["sucursal"], unidad_id))
            if not unidad_id:
                # La planilla de licencias tiene camionetas y unidades viejas
                # que la de gomería no tiene. Se dan de alta: son de la flota
                # igual, solo que no llevan mapa de cubiertas.
                resumen["unidades_nuevas"].append(f["patente"])
                if probar:
                    unidad_id = -1     # existiría; alcanza para el ensayo
                else:
                    fila = cx.execute("""
                        insert into unidades (patente, sucursal) values (%s,%s)
                        on conflict (patente) do update set patente = excluded.patente
                        returning id""", (f["patente"], f["sucursal"])).fetchone()
                    unidad_id = unidades[f["patente"]] = fila["id"]

        # --- la persona
        persona_id = None
        if f["chofer"]:
            fila = cx.execute("select id from personas where nombre = %s",
                              (f["chofer"],)).fetchone()
            if fila:
                persona_id = fila["id"]
                if not probar:
                    cx.execute("""update personas set sucursal = %s,
                                  unidad_id = coalesce(%s, unidad_id) where id = %s""",
                               (f["sucursal"], unidad_id, persona_id))
            else:
                resumen["personas_nuevas"] += 1
                if probar:
                    persona_id = -1
                else:
                    persona_id = cx.execute("""
                        insert into personas (nombre, sucursal, unidad_id)
                        values (%s,%s,%s) returning id""",
                        (f["chofer"], f["sucursal"], unidad_id)).fetchone()["id"]

        # --- los vencimientos
        for nombre_tipo, (alta, vence) in f["fechas"].items():
            tipo = tipos.get(nombre_tipo)
            if not tipo:
                continue
            de_unidad = tipo["ambito"] == "unidad"
            sujeto = unidad_id if de_unidad else persona_id
            if not sujeto:
                if de_unidad and f["chofer"]:
                    resumen["sin_patente"].append(f"{f['chofer']} · {nombre_tipo}")
                continue

            if probar:
                resumen["cargados"] += 1
                continue

            # Una fecha que ya está cargada no se duplica; una distinta
            # entra como renovación y la vista se queda con la más nueva.
            ya = cx.execute("""
                select 1 from vencimientos
                where tipo_id = %s and unidad_id is not distinct from %s
                  and persona_id is not distinct from %s and vence = %s""",
                (tipo["id"], unidad_id if de_unidad else None,
                 persona_id if not de_unidad else None, vence)).fetchone()
            if ya:
                resumen["ya_estaban"] += 1
                continue

            cx.execute("""
                insert into vencimientos
                  (tipo_id, unidad_id, persona_id, desde, vence, observaciones, usuario)
                values (%s,%s,%s,%s,%s,%s,'planilla')""",
                (tipo["id"], unidad_id if de_unidad else None,
                 persona_id if not de_unidad else None,
                 alta, vence, f["observaciones"]))
            resumen["cargados"] += 1

        resumen["sin_fecha"] += 4 - len(f["fechas"])

    return resumen


def main():
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    archivo = sys.argv[1]
    probar = "--probar" in sys.argv

    filas = leer(archivo)
    sucursales = {}
    for f in filas:
        sucursales[f["sucursal"]] = sucursales.get(f["sucursal"], 0) + 1
    print(f"{len(filas)} filas en {len(sucursales)} sucursales")
    for s, n in sucursales.items():
        print(f"  {s:18s} {n}")

    url = (os.environ.get("SUPABASE_DB_URL") or "").strip()
    if not url:
        raise SystemExit("\nFalta SUPABASE_DB_URL (la conexión a la base).")

    with psycopg.connect(url, row_factory=dict_row) as cx:
        r = importar(cx, filas, probar)
        if probar:
            cx.rollback()
        else:
            cx.commit()

    print(f"\n{'ASÍ QUEDARÍA (no se escribió nada)' if probar else 'LISTO'}")
    print(f"  vencimientos cargados : {r['cargados']}")
    print(f"  ya estaban            : {r['ya_estaban']}")
    print(f"  choferes nuevos       : {r['personas_nuevas']}")
    print(f"  fechas vacías         : {r['sin_fecha']}")
    nuevas = sorted(set(r["unidades_nuevas"]))
    if nuevas:
        print(f"  unidades que no estaban en la base ({len(nuevas)}):")
        print("    " + ", ".join(nuevas))
    if r["sin_patente"]:
        print(f"  sin dominio, no se pudo cargar la VTV ni el matafuego "
              f"({len(r['sin_patente'])}):")
        for x in r["sin_patente"][:10]:
            print("    " + x)


if __name__ == "__main__":
    main()
