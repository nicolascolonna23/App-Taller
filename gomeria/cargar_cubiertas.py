#!/usr/bin/env python3
"""
Carga el stock de cubiertas desde un CSV o un Excel.

    python3 gomeria/cargar_cubiertas.py stock.xlsx --simular
    python3 gomeria/cargar_cubiertas.py stock.xlsx

Columnas que entiende (el nombre no tiene que ser exacto: busca parecidos,
no importan mayúsculas ni acentos):

    codigo       obligatoria — número de fuego, serie, o el que usen
    marca        BRIDGESTONE, MICHELIN, PIRELLI...
    modelo
    medida       295/80R22.5
    estado       stock, montada, reparacion, recapado, baja. Por defecto stock.
    remanente    milímetros de dibujo
    km
    recapados    cuántas veces se recapó
    costo
    patente      si ya está puesta en una unidad
    posicion     1I, 2IE, 3DE... junto con patente, la monta

Si trae patente y posición, además de dar de alta la cubierta la monta en esa
unidad. Es la forma de cargar el inventario inicial de una sola vez.

Se puede correr las veces que haga falta: una cubierta que ya existe se
actualiza, no se duplica.
"""
import argparse, csv, os, re, sys, unicodedata

AQUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, AQUI)
import base

ESTADOS = ("stock", "montada", "reparacion", "recapado", "baja")

# Cada columna nuestra y las formas en que suele venir escrita.
ALIAS = {
    "codigo":    ["codigo", "cod", "nrofuego", "numerodefuego", "fuego", "serie", "id"],
    "marca":     ["marca"],
    "modelo":    ["modelo", "dibujo", "banda"],
    "medida":    ["medida", "rodado", "medidas", "tamano"],
    "estado":    ["estado", "situacion"],
    "remanente": ["remanente", "mm", "milimetros", "profundidad", "dibujomm"],
    "km":        ["km", "kilometros", "kmacumulados", "kmrecorridos"],
    "recapados": ["recapados", "recapado", "recauchutados"],
    "costo":     ["costo", "precio", "valor", "costocompra"],
    "patente":   ["patente", "dominio", "unidad", "movil"],
    "posicion":  ["posicion", "pos", "lugar", "ubicacion"],
    "nota":      ["nota", "observaciones", "obs", "comentario"],
}


def simplificar(t):
    """Saca acentos, espacios y signos: 'Nro. de Fuego' -> 'nrodefuego'."""
    t = unicodedata.normalize("NFKD", str(t or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", t.lower())


def mapear(encabezados):
    """Decide qué columna del archivo corresponde a cada campo nuestro.

    Tres pasadas, de la más exigente a la más suelta, para que 'Nro. de Fuego'
    caiga en codigo y 'Remanente (mm)' en remanente sin que haga falta
    renombrar nada en el archivo.
    """
    simples = {simplificar(h): h for h in encabezados if h}
    mapa, usadas = {}, set()

    def buscar(prueba):
        for s, original in simples.items():
            if original in usadas:
                continue
            if prueba(s):
                return s, original
        return None, None

    for pasada in ("exacta", "empieza", "contiene"):
        for campo, formas in ALIAS.items():
            if campo in mapa:
                continue
            if pasada == "exacta":
                prueba = lambda s, f=formas: s in f
            elif pasada == "empieza":
                prueba = lambda s, f=formas: any(s.startswith(x) for x in f)
            else:
                # 'medida' contiene 'id': por eso la pasada suelta solo vale
                # para alias largos, que no aparecen dentro de otra palabra.
                largas = [x for x in formas if len(x) >= 5]
                prueba = lambda s, f=largas: any(x in s for x in f)
            _, original = buscar(prueba)
            if original:
                mapa[campo] = original
                usadas.add(original)
    return mapa


def leer(ruta):
    """Devuelve (encabezados, filas) de un CSV o de un Excel."""
    if ruta.lower().endswith((".xlsx", ".xlsm", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        encabezados = [str(c).strip() if c is not None else "" for c in next(it)]
        filas = [dict(zip(encabezados, f)) for f in it]
        wb.close()
        return encabezados, filas
    with open(ruta, encoding="utf-8-sig", newline="") as f:
        lector = csv.DictReader(f)
        return list(lector.fieldnames or []), list(lector)


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".") if s.rfind(",") > s.rfind(".") else s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def texto(v):
    s = str(v).strip() if v is not None else ""
    return s or None


def main():
    ap = argparse.ArgumentParser(description="Carga el stock de cubiertas")
    ap.add_argument("archivo", help="CSV o Excel con el stock")
    ap.add_argument("--simular", action="store_true",
                    help="Muestra qué haría, sin escribir en la base")
    a = ap.parse_args()

    if not os.path.exists(a.archivo):
        raise SystemExit(f"No existe el archivo: {a.archivo}")

    encabezados, filas = leer(a.archivo)
    if not filas:
        raise SystemExit("El archivo está vacío.")

    mapa = mapear(encabezados)
    if "codigo" not in mapa:
        raise SystemExit(
            "No encontré la columna del código de la cubierta.\n"
            f"  Encabezados del archivo: {', '.join(str(h) for h in encabezados if h)}\n"
            "  Renombrá esa columna a 'codigo' y volvé a probar.")

    print(f"{len(filas)} filas en el archivo")
    print("Columnas reconocidas:")
    for campo in ALIAS:
        if campo in mapa:
            print(f"  {campo:<11} <- {mapa[campo]}")
    ignoradas = [h for h in encabezados if h and h not in mapa.values()]
    if ignoradas:
        print(f"  (se ignoran: {', '.join(str(h) for h in ignoradas)})")

    # Se valida todo antes de escribir nada.
    limpias, problemas, a_montar = [], [], 0
    vistos = set()
    for n, f in enumerate(filas, start=2):
        d = {campo: f.get(col) for campo, col in mapa.items()}
        codigo = texto(d.get("codigo"))
        if not codigo:
            continue
        if codigo.upper() in vistos:
            problemas.append(f"Línea {n}: el código {codigo} está repetido en el archivo.")
            continue
        vistos.add(codigo.upper())

        estado = (texto(d.get("estado")) or "stock").lower()
        estado = {"disponible": "stock", "en stock": "stock", "nueva": "stock",
                  "puesta": "montada", "en uso": "montada",
                  "reparación": "reparacion", "dada de baja": "baja"}.get(estado, estado)
        if estado not in ESTADOS:
            problemas.append(f"Línea {n} ({codigo}): estado '{estado}' desconocido. "
                             f"Tiene que ser uno de: {', '.join(ESTADOS)}.")
            continue

        pat, pos = texto(d.get("patente")), texto(d.get("posicion"))
        if pat and pos:
            a_montar += 1
        elif pat or pos:
            problemas.append(f"Línea {n} ({codigo}): para montarla necesito "
                             f"patente y posición, y falta {'la posición' if pat else 'la patente'}.")
            continue

        d.update(codigo=codigo, estado=estado, patente=pat, posicion=pos, _linea=n)
        limpias.append(d)

    if problemas:
        print(f"\n{len(problemas)} problemas:")
        for p in problemas[:20]:
            print(f"  {p}")
        if len(problemas) > 20:
            print(f"  ... y {len(problemas) - 20} más")
        raise SystemExit("\nNo cargué nada. Corregí el archivo y volvé a probar.")

    from collections import Counter
    print(f"\n{len(limpias)} cubiertas para cargar")
    for e, c in sorted(Counter(x["estado"] for x in limpias).items()):
        print(f"  {e:<12} {c}")
    if a_montar:
        print(f"  {a_montar} vienen con patente y posición: se montan al cargarlas")

    if a.simular:
        print("\n(simulación: no se escribió nada)")
        return

    nuevas = montadas = 0
    fallos = []
    with base.conectar() as cx:
        for d in limpias:
            cid = base.alta_cubierta(
                cx, d["codigo"], marca=texto(d.get("marca")), modelo=texto(d.get("modelo")),
                medida=texto(d.get("medida")), remanente_mm=num(d.get("remanente")),
                costo_compra=num(d.get("costo")), observaciones=texto(d.get("nota")),
                usuario="carga inicial")
            cx.execute("""update cubiertas set estado = %s,
                          km_acumulados = coalesce(%s, km_acumulados),
                          recapados = coalesce(%s, recapados)
                          where id = %s""",
                       (d["estado"] if not d["patente"] else "stock",
                        num(d.get("km")), int(num(d.get("recapados")) or 0), cid))
            nuevas += 1

            if d["patente"]:
                unidad = base.buscar_unidad(cx, d["patente"])
                if not unidad:
                    fallos.append(f"Línea {d['_linea']} ({d['codigo']}): "
                                  f"no existe la unidad {d['patente']}.")
                    continue
                posicion = base.posicion_por_codigo(cx, unidad["id"], d["posicion"])
                if not posicion:
                    fallos.append(f"Línea {d['_linea']} ({d['codigo']}): la posición "
                                  f"{d['posicion']} no existe en {d['patente']}.")
                    continue
                try:
                    base.montar(cx, unidad["id"], posicion["id"], cid,
                                km=unidad.get("km_actual"), usuario="carga inicial")
                    montadas += 1
                except Exception as e:
                    fallos.append(f"Línea {d['_linea']} ({d['codigo']}): {e}")
        cx.commit()

    print(f"\nCargadas {nuevas} cubiertas.")
    if montadas:
        print(f"Montadas {montadas} en sus unidades.")
    if fallos:
        print(f"\n{len(fallos)} no se pudieron montar (la cubierta sí quedó cargada):")
        for f in fallos[:20]:
            print(f"  {f}")
        if len(fallos) > 20:
            print(f"  ... y {len(fallos) - 20} más")


if __name__ == "__main__":
    main()
