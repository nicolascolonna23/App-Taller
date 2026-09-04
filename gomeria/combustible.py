"""
Cruce de remitos de combustible.  MÓDULO EN PRUEBA.

La estación de servicio manda un listado de remitos y después la factura.
Nosotros tenemos nuestra planilla de cargas. Hoy alguien compara las dos a
ojo antes de pagar. Acá se cruzan por número de remito y queda a la vista
lo que no coincide: lo que nos facturan y no tenemos, lo que cargamos y no
vino, y las diferencias de litros o de importe.

Está en prueba: se usa en paralelo con lo de siempre hasta que los números
den. Todo se carga por lotes, y un lote se borra entero.

Los archivos vienen del navegador en base64 y se leen acá, no allá: así el
mismo código lee el .xlsx de la estación y el .csv de la planilla, y no hay
que mantener dos parseos.
"""
import base64
import csv
import datetime
import io
import re

GESTORES = {"admin", "encargado"}

# Cómo se llama cada dato en los archivos que llegan. Cada estación arma su
# listado a su manera, así que se busca por lo que contiene el título, no
# por igualdad, y se prueban varios nombres.
ALIAS = {
    "remito":  ("REMITO", "COMPROBANTE", "TICKET", "NRO REMITO", "N REMITO", "VALE"),
    "fecha":   ("FECHA", "DIA"),
    "patente": ("PATENTE", "DOMINIO", "MOVIL", "MÓVIL", "UNIDAD", "CHAPA"),
    "litros":  ("LITROS", "LTS", "CANTIDAD", "VOLUMEN"),
    "importe": ("IMPORTE", "TOTAL", "MONTO", "PRECIO"),
    "estacion": ("ESTACION", "ESTACIÓN", "SURTIDOR", "PROVEEDOR", "RAZON SOCIAL"),
    "chofer":  ("CHOFER", "CONDUCTOR"),
}
# Cuando dos títulos entran en el mismo campo gana el primero de la lista,
# porque están ordenados de más específico a más general: "NRO REMITO" antes
# que "REMITO", y "TOTAL" no le puede ganar a "IMPORTE".


def _exigir_gestor(usuario, que="cargar combustible"):
    if (usuario or {}).get("rol") not in GESTORES:
        raise PermissionError(f"Solo un encargado o administrador puede {que}.")


def _remito(valor):
    """El número del remito, comparable venga como venga.

    La estación factura con el punto de venta adelante —0001-00123456— y
    nuestra planilla anota solo el número —123456—. Sacar todos los guiones
    y pegar los dígitos daría 100123456, que no es el mismo número: por eso
    lo que va antes del guión se descarta, que es lo que significa.

    0001-00123456, 00123456, 123456 y R 123.456 quedan todos en 123456.
    """
    texto = str(valor or "")
    if "-" in texto:
        texto = texto.rsplit("-", 1)[1]
    return re.sub(r"[^0-9]", "", texto).lstrip("0") or ""


def _patente(valor):
    return re.sub(r"[^A-Z0-9]", "", str(valor or "").upper()) or None


def _numero(valor):
    """Números en formato argentino y en el crudo del Excel."""
    if isinstance(valor, (int, float)):
        return None if valor != valor else float(valor)
    s = re.sub(r"[\s$]", "", str(valor or ""))
    if not s or not re.fullmatch(r"[\d.,-]+", s):
        return None
    coma, punto = s.rfind(","), s.rfind(".")
    if coma >= 0 and punto >= 0:
        dec = "," if coma > punto else "."
        s = s.replace("." if dec == "," else ",", "").replace(dec, ".")
    elif coma >= 0:
        s = s.replace(",", "") if re.fullmatch(r"\d{1,3}(,\d{3})+", s) else s.replace(",", ".")
    elif punto >= 0 and re.fullmatch(r"\d{1,3}(\.\d{3})+", s):
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return None


def _fecha(valor):
    if isinstance(valor, datetime.datetime):
        return valor.date()
    if isinstance(valor, datetime.date):
        return valor
    texto = str(valor or "").strip()[:10]
    m = re.match(r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})$", texto)
    if m:
        anio = m[3] if len(m[3]) == 4 else "20" + m[3]
        try:
            return datetime.date(int(anio), int(m[2]), int(m[1]))
        except ValueError:
            return None
    try:
        return datetime.date.fromisoformat(texto)
    except ValueError:
        return None


# =====================================================================
# LEER EL ARCHIVO
# =====================================================================
def _filas_de(nombre, crudo):
    """Las filas del archivo, sea .xlsx o .csv, como listas de celdas."""
    if nombre.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        libro = openpyxl.load_workbook(io.BytesIO(crudo), data_only=True, read_only=True)
        hoja = libro[libro.sheetnames[0]]
        return [list(f) for f in hoja.iter_rows(values_only=True)]

    texto = crudo.decode("utf-8-sig", "replace")
    # La coma y el punto y coma se usan las dos; gana la que más aparece en
    # la primera línea, que es donde están los títulos.
    primera = texto.split("\n", 1)[0]
    sep = ";" if primera.count(";") > primera.count(",") else ","
    return [f for f in csv.reader(io.StringIO(texto), delimiter=sep)]


def _cabecera(filas, elegidas=None):
    """Dónde están los títulos y qué columna es cada cosa.

    Adivina por el nombre del título, pero adivinar no alcanza: una planilla
    de cargas puede tener a la vez el número de ticket y el de remito, y el
    que sirve para cruzar es el que figura en la factura. Por eso `elegidas`
    manda por encima de todo: es lo que el usuario eligió en la pantalla,
    columna por columna.
    """
    elegidas = elegidas or {}
    for i, fila in enumerate(filas[:15]):
        titulos = [" ".join(str(c or "").upper().split()) for c in fila]
        if not any(titulos):
            continue
        indice = {}
        for campo, nombres in ALIAS.items():
            # Lo elegido a mano gana: se busca por el nombre exacto de la
            # columna, que es lo que la pantalla devuelve.
            if campo in elegidas:
                j = next((k for k, t in enumerate(titulos)
                          if t == elegidas[campo]), None)
                if j is not None:
                    indice[campo] = j
                    continue
            for n in nombres:
                j = next((k for k, t in enumerate(titulos)
                          if n in t and k not in indice.values()), None)
                if j is not None:
                    indice[campo] = j
                    break
        if "remito" in indice:
            return i, indice, titulos
    return None, None, None


def leer(nombre, crudo, elegidas=None):
    """El archivo hecho filas listas para guardar, más lo que se descartó."""
    filas = _filas_de(nombre, crudo)
    if not filas:
        raise ValueError("El archivo está vacío.")

    fila_titulos, indice, titulos = _cabecera(filas, elegidas)
    if indice is None:
        vistos = ", ".join(str(c) for c in filas[0][:8] if c)
        raise ValueError(
            "No encuentro la columna del remito. Tiene que haber una que se "
            f"llame REMITO, COMPROBANTE, TICKET o VALE. En la primera fila vi: {vistos}")

    def celda(fila, campo):
        j = indice.get(campo)
        return fila[j] if j is not None and j < len(fila) else None

    salida, descartadas = [], 0
    for fila in filas[fila_titulos + 1:]:
        remito = _remito(celda(fila, "remito"))
        if not remito:
            descartadas += 1          # totales, subtotales, filas en blanco
            continue
        salida.append({
            "remito": remito,
            "remito_bruto": str(celda(fila, "remito") or "").strip()[:40],
            "fecha": _fecha(celda(fila, "fecha")),
            "patente": _patente(celda(fila, "patente")),
            "litros": _numero(celda(fila, "litros")),
            "importe": _numero(celda(fila, "importe")),
            "estacion": (str(celda(fila, "estacion") or "").strip() or None),
            "chofer": (str(celda(fila, "chofer") or "").strip().upper() or None),
        })

    # Un remito repetido adentro del mismo archivo es un error de armado y
    # hay que verlo antes de guardar, no después.
    repetidos = {}
    for f in salida:
        repetidos[f["remito"]] = repetidos.get(f["remito"], 0) + 1
    repetidos = sorted(r for r, n in repetidos.items() if n > 1)

    return {"filas": salida, "descartadas": descartadas, "repetidos": repetidos,
            "columnas": sorted(indice),
            # Qué columna terminó siendo cada cosa, y todas las que hay:
            # con eso la pantalla arma los selectores para corregirlo.
            "usadas": {campo: titulos[j] for campo, j in indice.items()},
            "cabeceras": [t for t in titulos if t]}


# =====================================================================
# GUARDAR
# =====================================================================
def subir(cx, datos, usuario=None):
    """Lee el archivo y, si `confirmar` viene, lo guarda. Devuelve el resumen."""
    _exigir_gestor(usuario)

    origen = (datos.get("origen") or "").strip()
    if origen not in ("estacion", "planilla"):
        raise ValueError("Decí si el archivo es el listado de la estación o nuestra planilla.")
    nombre = (datos.get("nombre") or "archivo").strip()[:120]
    try:
        crudo = base64.b64decode(datos.get("contenido") or "", validate=False)
    except Exception:
        raise ValueError("El archivo no se pudo leer.")
    if not crudo:
        raise ValueError("El archivo llegó vacío.")

    leido = leer(nombre, crudo, datos.get("columnas"))
    filas = leido["filas"]
    if not filas:
        raise ValueError("No encontré ninguna fila con número de remito.")

    # Sin confirmar solo se muestra qué entraría. La primera vez conviene
    # mirarlo: si la estación cambió el formato, se ve acá y no después de
    # haber ensuciado la tabla.
    if not datos.get("confirmar"):
        return {"previo": True, "leidas": len(filas),
                "descartadas": leido["descartadas"],
                "repetidos": leido["repetidos"],
                "columnas": leido["columnas"],
                "usadas": leido["usadas"],
                "cabeceras": leido["cabeceras"],
                "muestra": filas[:8]}

    lote = cx.execute("""
        insert into combustible_lotes (origen, archivo, estacion, periodo, filas, usuario)
        values (%s,%s,%s,%s,%s,%s) returning id""",
        (origen, nombre,
         (datos.get("estacion") or "").strip()[:120] or None,
         (datos.get("periodo") or "").strip()[:20] or None,
         len(filas), (usuario or {}).get("nombre"))).fetchone()["id"]

    # El remito que ya estaba se pisa: subir de nuevo el listado corregido
    # tiene que dejar la última versión, no dos.
    antes = cx.execute("select count(*) as n from combustible_cargas where origen = %s",
                       (origen,)).fetchone()["n"]
    for f in filas:
        cx.execute("""
            insert into combustible_cargas
              (lote_id, origen, remito, remito_bruto, fecha, patente,
               litros, importe, estacion, chofer)
            values (%(lote)s,%(origen)s,%(remito)s,%(bruto)s,%(fecha)s,%(patente)s,
                    %(litros)s,%(importe)s,%(estacion)s,%(chofer)s)
            on conflict (origen, remito, coalesce(patente, '')) do update set
              lote_id = excluded.lote_id, fecha = excluded.fecha,
              patente = excluded.patente, litros = excluded.litros,
              importe = excluded.importe, estacion = excluded.estacion,
              chofer = excluded.chofer, remito_bruto = excluded.remito_bruto""",
            {"lote": lote, "origen": origen, "remito": f["remito"],
             "bruto": f["remito_bruto"], "fecha": f["fecha"], "patente": f["patente"],
             "litros": f["litros"], "importe": f["importe"],
             "estacion": f["estacion"] or (datos.get("estacion") or None),
             "chofer": f["chofer"]})
    despues = cx.execute("select count(*) as n from combustible_cargas where origen = %s",
                         (origen,)).fetchone()["n"]

    return {"previo": False, "lote": lote, "leidas": len(filas),
            "repetidos": leido["repetidos"],
            "nuevas": despues - antes, "actualizadas": len(filas) - (despues - antes),
            "descartadas": leido["descartadas"]}


def borrar_lote(cx, lote_id, usuario=None):
    _exigir_gestor(usuario, "borrar una carga")
    fila = cx.execute("delete from combustible_lotes where id = %s returning archivo",
                      (lote_id,)).fetchone()
    if not fila:
        raise ValueError("Ese lote no existe.")
    return {"borrado": fila["archivo"]}


# =====================================================================
# LEER
# =====================================================================
def _uno(cx, consulta, valores=()):
    try:
        return cx.execute(consulta, valores).fetchall()
    except Exception:
        cx.rollback()
        return []


def panel(cx, estado=None, limite=400):
    """El cruce, el resumen y los lotes cargados."""
    filtro, valores = "", []
    if estado:
        filtro = "where estado = %s"
        valores.append(estado)
    valores.append(limite)
    return {
        "resumen": _uno(cx, "select * from v_combustible_resumen order by estado"),
        "cruce": _uno(cx, f"""
            select c.*, u.interno, u.chofer as chofer_unidad
            from v_combustible_cruce c
            left join unidades u on u.id = c.unidad_id
            {filtro}
            -- Primero lo que hay que mirar y después lo que está bien.
            order by (c.estado = 'ok'), c.fecha desc nulls last, c.remito
            limit %s""", valores),
        "lotes": _uno(cx, """
            select l.*, count(c.id)::int as vigentes
            from combustible_lotes l
            left join combustible_cargas c on c.lote_id = l.id
            group by l.id order by l.subido desc limit 30"""),
    }
