#!/usr/bin/env python3
"""
Ingesta de los reportes a una base SQLite.

Lee los .xlsx (de disco o directamente del servidor de BI) y arma datos.db,
que es contra lo que consulta el chat. Se corre cada vez que hay reportes
nuevos; es idempotente: reescribe las tablas de cero.

    python3 ingesta.py                          # usa las URLs de config.json
    python3 ingesta.py --clientes ruta.xlsx --cuenta-corriente ruta.xlsx

Bajar los xlsx desde acá (y no desde el navegador) es lo que resuelve el
problema de CORS: bi.sistemaexpreso.com.ar no le contesta a una página web,
pero sí a un proceso del servidor.
"""
import argparse, json, os, sqlite3, sys, tempfile, urllib.request
from datetime import datetime

AQUI = os.path.dirname(os.path.abspath(__file__))
DB   = os.path.join(AQUI, "datos.db")
CONF = os.path.join(AQUI, "config.json")


def cargar_config():
    if os.path.exists(CONF):
        with open(CONF, encoding="utf-8") as f:
            return json.load(f)
    return {}


def obtener(origen):
    """Devuelve una ruta local. Si origen es una URL, la baja a un temporal."""
    if not origen:
        return None
    if origen.startswith(("http://", "https://")):
        print(f"  bajando {origen}")
        destino = os.path.join(tempfile.gettempdir(), os.path.basename(origen))
        req = urllib.request.Request(origen, headers={"User-Agent": "ingesta-diemar/1.0"})
        with urllib.request.urlopen(req, timeout=300) as r, open(destino, "wb") as f:
            f.write(r.read())
        return destino
    if not os.path.exists(origen):
        raise SystemExit(f"No existe el archivo: {origen}")
    return origen


def filas(ruta):
    """Itera un xlsx como dicts {encabezado: valor}, sin cargarlo entero en memoria."""
    import openpyxl
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    encabezados = [str(c).strip() if c is not None else "" for c in next(it)]
    for fila in it:
        yield dict(zip(encabezados, fila))
    wb.close()


def num(v):
    """Number tolerante: '1.383.254,30' y '1383254.3' dan lo mismo."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(" ", "")
    if not s:
        return None
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".") if s.rfind(",") > s.rfind(".") else s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def fecha(v):
    """Normaliza a AAAA-MM-DD para poder ordenar y comparar en SQL."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()[:10]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def texto(v):
    return None if v is None else str(v).strip()


ESQUEMA = """
DROP TABLE IF EXISTS clientes;
CREATE TABLE clientes (
  id INTEGER PRIMARY KEY, razon_social TEXT, alias TEXT, nif TEXT,
  telefono TEXT, email TEXT, provincia TEXT, localidad TEXT, direccion TEXT,
  codigo_postal TEXT, tarifario TEXT, condicion_pago TEXT, sucursal TEXT,
  limite_credito REAL, contacto_cobro TEXT, contacto_cobro_email TEXT,
  contacto_cobro_telefono TEXT, condicion_iva TEXT, vto_poliza TEXT, monto_poliza REAL
);
DROP TABLE IF EXISTS cuenta_corriente;
CREATE TABLE cuenta_corriente (
  id_cliente INTEGER, nif TEXT, estado TEXT, empresa TEXT, nro TEXT, tipo TEXT,
  fecha_emision TEXT, fecha_vencimiento TEXT, concepto TEXT,
  monto REAL, neto REAL, iva REAL, monto_pendiente REAL, vencido INTEGER
);
"""

INDICES = """
CREATE INDEX ix_cli_razon    ON clientes(razon_social);
CREATE INDEX ix_cli_nif      ON clientes(nif);
CREATE INDEX ix_cli_sucursal ON clientes(sucursal);
CREATE INDEX ix_cli_prov     ON clientes(provincia);
CREATE INDEX ix_cc_cliente   ON cuenta_corriente(id_cliente);
CREATE INDEX ix_cc_tipo      ON cuenta_corriente(tipo);
CREATE INDEX ix_cc_estado    ON cuenta_corriente(estado);
CREATE INDEX ix_cc_emision   ON cuenta_corriente(fecha_emision);
"""


def ingerir_clientes(cx, ruta):
    n = 0
    for f in filas(ruta):
        try:
            idc = int(str(f.get("Id") or "").strip())
        except (ValueError, TypeError):
            continue
        cx.execute(
            "INSERT OR REPLACE INTO clientes VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (idc, texto(f.get("Razon Social")), texto(f.get("Alias")), texto(f.get("Nif")),
             texto(f.get("Teléfono")), texto(f.get("Email")), texto(f.get("Provincia")),
             texto(f.get("Localidad")), texto(f.get("Dirección")), texto(f.get("Código Postal")),
             texto(f.get("Tarifario")), texto(f.get("Condición de Pago")), texto(f.get("Sucursal")),
             num(f.get("Límite Crédito")), texto(f.get("Contacto Cobro")),
             texto(f.get("Contacto Cobro Email")), texto(f.get("Contacto Cobro Teléfono")),
             texto(f.get("Condición IVA")), fecha(f.get("VTO. Póliza")), num(f.get("Monto Póliza"))))
        n += 1
    return n


def ingerir_cuenta_corriente(cx, ruta):
    n = 0
    lote = []
    for f in filas(ruta):
        try:
            idc = int(str(f.get("Id Cliente") or "").strip())
        except (ValueError, TypeError):
            continue
        lote.append((idc, texto(f.get("Nif")), texto(f.get("Estado")), texto(f.get("Empresa")),
                     texto(f.get("Nro")), texto(f.get("Tipo")), fecha(f.get("Fecha de Emisión")),
                     fecha(f.get("Fecha de Vencimiento")), texto(f.get("Concepto")),
                     num(f.get("Monto")), num(f.get("Neto")), num(f.get("Iva")),
                     num(f.get("Monto Pendiente")) or 0.0,
                     1 if str(f.get("Vencido") or "").strip() == "1" else 0))
        n += 1
        if len(lote) >= 5000:
            cx.executemany("INSERT INTO cuenta_corriente VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)", lote)
            lote = []
    if lote:
        cx.executemany("INSERT INTO cuenta_corriente VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)", lote)
    return n


def main():
    conf = cargar_config()
    ap = argparse.ArgumentParser(description="Arma datos.db a partir de los reportes")
    ap.add_argument("--clientes", default=conf.get("origen_clientes"))
    ap.add_argument("--cuenta-corriente", default=conf.get("origen_cuenta_corriente"))
    a = ap.parse_args()

    if not a.clientes and not a.cuenta_corriente:
        raise SystemExit("Indicá --clientes y/o --cuenta-corriente, o completá config.json")

    print("Ingesta de reportes")
    cx = sqlite3.connect(DB)
    cx.executescript(ESQUEMA)

    if a.clientes:
        print("Clientes:")
        n = ingerir_clientes(cx, obtener(a.clientes))
        print(f"  {n:,} clientes")
    if a.cuenta_corriente:
        print("Cuenta corriente:")
        n = ingerir_cuenta_corriente(cx, obtener(a.cuenta_corriente))
        print(f"  {n:,} movimientos")

    cx.executescript(INDICES)
    cx.commit()
    cx.execute("ANALYZE")
    cx.commit()
    cx.close()
    print(f"\nListo: {DB} ({os.path.getsize(DB)/1e6:.1f} MB)")


if __name__ == "__main__":
    main()
